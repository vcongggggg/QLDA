from io import BytesIO
from datetime import datetime, timezone
import csv

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


def build_kpi_csv(rows: list[dict]) -> str:
    lines = ["user_id,user_name,month,done_on_time,done_late,overdue_unfinished,score,target_score,progress_percent,gap"]
    for r in rows:
        lines.append(
            f"{r['user_id']},{r['user_name']},{r['month']},{r['done_on_time']},{r['done_late']},{r['overdue_unfinished']},{r['score']},{r.get('target_score', '')},{r.get('progress_percent', '')},{r.get('gap', '')}"
        )
    return "\n".join(lines)


def build_kpi_xlsx(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI"
    ws.append(["user_id", "user_name", "month", "done_on_time", "done_late", "overdue_unfinished", "score", "target_score", "progress_percent", "gap"])
    for r in rows:
        ws.append(
            [
                r["user_id"],
                r["user_name"],
                r["month"],
                r["done_on_time"],
                r["done_late"],
                r["overdue_unfinished"],
                r["score"],
                r.get("target_score"),
                r.get("progress_percent"),
                r.get("gap"),
            ]
        )
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_kpi_pdf(rows: list[dict], month: str) -> bytes:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 50

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(40, y, f"TeamsWork KPI Report - {month}")
    y -= 24
    pdf.setFont("Helvetica", 9)
    pdf.drawString(40, y, f"Generated at: {datetime.utcnow().isoformat()}Z")
    y -= 24
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(40, y, "User")
    pdf.drawString(220, y, "On-time")
    pdf.drawString(280, y, "Late")
    pdf.drawString(330, y, "Overdue")
    pdf.drawString(400, y, "Score")
    pdf.drawString(455, y, "Target")
    y -= 16
    pdf.setFont("Helvetica", 9)

    for row in rows:
        if y < 50:
            pdf.showPage()
            y = height - 50
            pdf.setFont("Helvetica", 9)
        pdf.drawString(40, y, str(row["user_name"])[:28])
        pdf.drawString(220, y, str(row["done_on_time"]))
        pdf.drawString(280, y, str(row["done_late"]))
        pdf.drawString(330, y, str(row["overdue_unfinished"]))
        pdf.drawString(400, y, str(row["score"]))
        pdf.drawString(455, y, str(row.get("target_score", "")))
        y -= 14

    pdf.save()
    return buffer.getvalue()


def build_project_progress_csv(rows: list[dict]) -> str:
    lines = [
        "project_id,total_tasks,done_tasks,overdue_tasks,completion_rate,total_story_points,completed_story_points"
    ]
    for r in rows:
        lines.append(
            f"{r['project_id']},{r['total_tasks']},{r['done_tasks']},{r['overdue_tasks']},{r['completion_rate']},{r['total_story_points']},{r['completed_story_points']}"
        )
    return "\n".join(lines)


def build_project_progress_xlsx(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "ProjectProgress"
    ws.append(
        [
            "project_id",
            "total_tasks",
            "done_tasks",
            "overdue_tasks",
            "completion_rate",
            "total_story_points",
            "completed_story_points",
        ]
    )
    for r in rows:
        ws.append(
            [
                r["project_id"],
                r["total_tasks"],
                r["done_tasks"],
                r["overdue_tasks"],
                r["completion_rate"],
                r["total_story_points"],
                r["completed_story_points"],
            ]
        )
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _parse_dt(value: object) -> datetime | None:
    if value in (None, ""):
        return None
    try:
        text = str(value).replace("Z", "+00:00")
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def build_report_analytics_summary(
    tasks: list[dict],
    *,
    month: str,
    as_of: datetime | None = None,
    generated_at: datetime | None = None,
    dependency_edges: list[dict] | None = None,
) -> dict:
    now = generated_at or datetime.now(timezone.utc)
    if now.tzinfo is None:
        now = now.replace(tzinfo=timezone.utc)
    cutoff = as_of or now
    if cutoff.tzinfo is None:
        cutoff = cutoff.replace(tzinfo=timezone.utc)
    cutoff = cutoff.astimezone(timezone.utc)

    rows = [
        task
        for task in tasks
        if str(task.get("deadline") or "")[:7] == month or str(task.get("completed_at") or "")[:7] == month
    ]
    total_tasks = len(rows)
    done_tasks = sum(1 for task in rows if task.get("status") == "done")
    open_tasks = total_tasks - done_tasks
    todo_tasks = sum(1 for task in rows if task.get("status") == "todo")
    doing_tasks = sum(1 for task in rows if task.get("status") == "doing")
    total_story_points = sum(int(task.get("story_points") or 0) for task in rows)
    completed_story_points = sum(int(task.get("story_points") or 0) for task in rows if task.get("status") == "done")
    assigned_tasks = sum(1 for task in rows if task.get("assignee_id") is not None)
    unassigned_tasks = total_tasks - assigned_tasks
    assigned_story_points = sum(int(task.get("story_points") or 0) for task in rows if task.get("assignee_id") is not None)
    unassigned_story_points = total_story_points - assigned_story_points

    def is_overdue_open(task: dict) -> bool:
        deadline = _parse_dt(task.get("deadline"))
        return bool(task.get("status") != "done" and deadline and deadline < cutoff)

    cycle_days: list[float] = []
    for task in rows:
        if task.get("status") != "done":
            continue
        created_at = _parse_dt(task.get("created_at"))
        completed_at = _parse_dt(task.get("completed_at"))
        if created_at and completed_at and completed_at >= created_at:
            cycle_days.append(round((completed_at - created_at).total_seconds() / 86400, 2))

    workload: dict[int | None, dict] = {}
    velocity: dict[int | None, dict] = {}
    project_effort: dict[int | None, dict] = {}
    for task in rows:
        user_id = int(task["assignee_id"]) if task.get("assignee_id") is not None else None
        item = workload.setdefault(
            user_id,
            {
                "user_id": user_id,
                "assignee_name": task.get("assignee_name") or "Unassigned",
                "total_tasks": 0,
                "done_tasks": 0,
                "open_tasks": 0,
                "overdue_open_tasks": 0,
                "story_points": 0,
                "completed_story_points": 0,
            },
        )
        points = int(task.get("story_points") or 0)
        item["total_tasks"] += 1
        item["story_points"] += points
        if task.get("status") == "done":
            item["done_tasks"] += 1
            item["completed_story_points"] += points
        else:
            item["open_tasks"] += 1
            if is_overdue_open(task):
                item["overdue_open_tasks"] += 1

        sprint_id = int(task["sprint_id"]) if task.get("sprint_id") is not None else None
        sprint_item = velocity.setdefault(
            sprint_id,
            {
                "sprint_id": sprint_id,
                "sprint_name": task.get("sprint_name") or ("Backlog" if sprint_id is None else f"Sprint {sprint_id}"),
                "planned_story_points": 0,
                "completed_story_points": 0,
            },
        )
        sprint_item["planned_story_points"] += points
        if task.get("status") == "done":
            sprint_item["completed_story_points"] += points

        project_id = int(task["project_id"]) if task.get("project_id") is not None else None
        project_item = project_effort.setdefault(
            project_id,
            {
                "project_id": project_id,
                "project_name": task.get("project_name") or ("No project" if project_id is None else f"Project {project_id}"),
                "total_tasks": 0,
                "done_tasks": 0,
                "overdue_open_tasks": 0,
                "story_points": 0,
                "completed_story_points": 0,
            },
        )
        project_item["total_tasks"] += 1
        project_item["story_points"] += points
        if task.get("status") == "done":
            project_item["done_tasks"] += 1
            project_item["completed_story_points"] += points
        elif is_overdue_open(task):
            project_item["overdue_open_tasks"] += 1

    velocity_rows = []
    for item in velocity.values():
        planned = int(item["planned_story_points"])
        completed = int(item["completed_story_points"])
        velocity_rows.append({**item, "completion_rate": round((completed / planned) * 100, 2) if planned else 0.0})

    project_rows = []
    for item in project_effort.values():
        total = int(item["total_tasks"])
        done = int(item["done_tasks"])
        project_rows.append({**item, "completion_rate": round((done / total) * 100, 2) if total else 0.0})

    dependency_edges = dependency_edges or []
    month_task_ids = {int(task["id"]) for task in rows if task.get("id") is not None}
    blocked_task_ids = {int(edge["task_id"]) for edge in dependency_edges if int(edge["task_id"]) in month_task_ids}
    source_task_ids = {
        int(edge["dependency_task_id"])
        for edge in dependency_edges
        if int(edge["task_id"]) in month_task_ids or int(edge["dependency_task_id"]) in month_task_ids
    }

    return {
        "month": month,
        "scope": {"type": "all", "user_id": None},
        "generated_at": now.astimezone(timezone.utc).isoformat(),
        "productivity": {
            "total_tasks": total_tasks,
            "done_tasks": done_tasks,
            "open_tasks": open_tasks,
            "completion_rate": round((done_tasks / total_tasks) * 100, 2) if total_tasks else 0.0,
            "total_story_points": total_story_points,
            "completed_story_points": completed_story_points,
        },
        "workload_distribution": sorted(
            workload.values(),
            key=lambda item: (-int(item["open_tasks"]), -int(item["total_tasks"]), str(item["assignee_name"])),
        ),
        "backlog_health": {
            "overdue_open_tasks": sum(1 for task in rows if is_overdue_open(task)),
            "backlog_open_tasks": sum(1 for task in rows if task.get("status") != "done" and task.get("sprint_id") is None),
            "unassigned_open_tasks": sum(1 for task in rows if task.get("status") != "done" and task.get("assignee_id") is None),
        },
        "cycle_time": {
            "done_tasks_with_cycle_time": len(cycle_days),
            "avg_cycle_time_days": round(sum(cycle_days) / len(cycle_days), 2) if cycle_days else None,
            "min_cycle_time_days": min(cycle_days) if cycle_days else None,
            "max_cycle_time_days": max(cycle_days) if cycle_days else None,
        },
        "task_status": {
            "todo_tasks": todo_tasks,
            "doing_tasks": doing_tasks,
            "done_tasks": done_tasks,
        },
        "utilization": {
            "assigned_tasks": assigned_tasks,
            "unassigned_tasks": unassigned_tasks,
            "assigned_story_points": assigned_story_points,
            "unassigned_story_points": unassigned_story_points,
            "utilization_rate": round((assigned_tasks / total_tasks) * 100, 2) if total_tasks else 0.0,
        },
        "velocity": sorted(
            velocity_rows,
            key=lambda item: (item["sprint_id"] is None, item["sprint_id"] or 0, str(item["sprint_name"])),
        ),
        "project_effort": sorted(
            project_rows,
            key=lambda item: (-int(item["story_points"]), -int(item["total_tasks"]), str(item["project_name"])),
        ),
        "dependency_map": {
            "total_dependency_edges": sum(
                1 for edge in dependency_edges if int(edge["task_id"]) in month_task_ids
            ),
            "blocked_tasks": len(blocked_task_ids),
            "dependency_source_tasks": len(source_task_ids),
        },
    }


def _analytics_export_rows(summary: dict) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []

    def add(section: str, metric: str, value: object, label: str = "") -> None:
        rows.append({"section": section, "metric": metric, "label": label, "value": value})

    productivity = summary.get("productivity") or {}
    for key in ("total_tasks", "done_tasks", "open_tasks", "completion_rate", "total_story_points", "completed_story_points"):
        add("productivity", key, productivity.get(key, ""))
    backlog = summary.get("backlog_health") or {}
    for key in ("overdue_open_tasks", "backlog_open_tasks", "unassigned_open_tasks"):
        add("backlog_health", key, backlog.get(key, ""))
    cycle = summary.get("cycle_time") or {}
    for key in ("done_tasks_with_cycle_time", "avg_cycle_time_days", "min_cycle_time_days", "max_cycle_time_days"):
        add("cycle_time", key, cycle.get(key, ""))
    utilization = summary.get("utilization") or {}
    for key in ("assigned_tasks", "unassigned_tasks", "assigned_story_points", "unassigned_story_points", "utilization_rate"):
        add("utilization", key, utilization.get(key, ""))
    dependency = summary.get("dependency_map") or {}
    for key in ("total_dependency_edges", "blocked_tasks", "dependency_source_tasks"):
        add("dependency_map", key, dependency.get(key, ""))
    for row in summary.get("workload_distribution") or []:
        add("workload", "open_tasks", row.get("open_tasks", ""), row.get("assignee_name", ""))
        add("workload", "done_tasks", row.get("done_tasks", ""), row.get("assignee_name", ""))
        add("workload", "story_points", row.get("story_points", ""), row.get("assignee_name", ""))
    for row in summary.get("velocity") or []:
        add("velocity", "planned_story_points", row.get("planned_story_points", ""), row.get("sprint_name", ""))
        add("velocity", "completed_story_points", row.get("completed_story_points", ""), row.get("sprint_name", ""))
    for row in summary.get("project_effort") or []:
        add("project_effort", "story_points", row.get("story_points", ""), row.get("project_name", ""))
        add("project_effort", "completed_story_points", row.get("completed_story_points", ""), row.get("project_name", ""))
    return rows


def build_report_analytics_csv(summary: dict) -> str:
    from io import StringIO

    headers = ["month", "generated_at", "section", "metric", "label", "value"]
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers)
    writer.writeheader()
    for row in _analytics_export_rows(summary):
        writer.writerow({"month": summary.get("month"), "generated_at": summary.get("generated_at"), **row})
    return buffer.getvalue()


def build_report_analytics_xlsx(summary: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics"
    ws.append(["month", "generated_at", "section", "metric", "label", "value"])
    for row in _analytics_export_rows(summary):
        ws.append([summary.get("month"), summary.get("generated_at"), row["section"], row["metric"], row["label"], row["value"]])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_rbac_matrix_csv(matrix: dict) -> str:
    roles = matrix["roles"]
    permissions = matrix["permissions"]
    role_keys = [str(role["slug"]) for role in roles]
    lines = ["permission_key,category,name," + ",".join(role_keys)]
    for permission in permissions:
        row = [
            str(permission["key"]),
            str(permission["category"]),
            str(permission["name"]).replace(",", " "),
        ]
        for role_key in role_keys:
            granted = str(permission["key"]) in set(matrix["matrix"].get(role_key, []))
            row.append("yes" if granted else "")
        lines.append(",".join(row))
    return "\n".join(lines)


def build_rbac_matrix_xlsx(matrix: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "RBAC Matrix"
    roles = matrix["roles"]
    permissions = matrix["permissions"]
    role_keys = [str(role["slug"]) for role in roles]
    ws.append(["permission_key", "category", "name", *role_keys])
    for permission in permissions:
        row = [permission["key"], permission["category"], permission["name"]]
        for role_key in role_keys:
            granted = str(permission["key"]) in set(matrix["matrix"].get(role_key, []))
            row.append("yes" if granted else "")
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_tasks_csv(rows: list[dict]) -> str:
    from io import StringIO

    headers = [
        "id",
        "title",
        "description",
        "assignee_id",
        "project_id",
        "sprint_id",
        "status",
        "story_points",
        "difficulty",
        "priority",
        "labels",
        "deadline",
        "completed_at",
    ]
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        item = {key: row.get(key) for key in headers}
        item["labels"] = ";".join(row.get("labels") or [])
        writer.writerow(item)
    return buffer.getvalue()


def build_tasks_xlsx(rows: list[dict]) -> bytes:
    headers = [
        "id",
        "title",
        "description",
        "assignee_id",
        "project_id",
        "sprint_id",
        "status",
        "story_points",
        "difficulty",
        "priority",
        "labels",
        "deadline",
        "completed_at",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.append(headers)
    for row in rows:
        ws.append([
            ";".join(row.get("labels") or []) if key == "labels" else row.get(key)
            for key in headers
        ])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_sprint_review_csv(summary: dict) -> str:
    headers = [
        "sprint_id",
        "project_id",
        "sprint_name",
        "status",
        "total_tasks",
        "done_tasks",
        "unfinished_tasks",
        "planned_story_points",
        "completed_story_points",
        "completion_rate",
    ]
    values = [str(summary.get(h, "")) for h in headers]
    return ",".join(headers) + "\n" + ",".join(values)


def build_sprint_review_xlsx(summary: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "SprintReview"
    for k, v in summary.items():
        ws.append([k, v])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_portfolio_csv(rows: list[dict]) -> str:
    headers = [
        "project_id",
        "project_name",
        "status",
        "completion_rate",
        "total_tasks",
        "overdue_tasks",
        "completed_story_points",
        "total_story_points",
    ]
    lines = [",".join(headers)]
    for r in rows:
        lines.append(
            ",".join(
                [
                    str(r.get("project_id", "")),
                    str(r.get("project_name", "")),
                    str(r.get("status", "")),
                    str(r.get("completion_rate", "")),
                    str(r.get("total_tasks", "")),
                    str(r.get("overdue_tasks", "")),
                    str(r.get("completed_story_points", "")),
                    str(r.get("total_story_points", "")),
                ]
            )
        )
    return "\n".join(lines)


def build_portfolio_xlsx(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"
    ws.append(
        [
            "project_id",
            "project_name",
            "status",
            "completion_rate",
            "total_tasks",
            "overdue_tasks",
            "completed_story_points",
            "total_story_points",
        ]
    )
    for r in rows:
        ws.append(
            [
                r.get("project_id"),
                r.get("project_name"),
                r.get("status"),
                r.get("completion_rate"),
                r.get("total_tasks"),
                r.get("overdue_tasks"),
                r.get("completed_story_points"),
                r.get("total_story_points"),
            ]
        )
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_audit_csv(rows: list[dict]) -> str:
    from io import StringIO

    headers = ["id", "actor_user_id", "actor_name", "action", "entity", "entity_id", "detail", "created_at"]
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key) for key in headers})
    return buffer.getvalue()


def build_audit_xlsx(rows: list[dict]) -> bytes:
    headers = ["id", "actor_user_id", "actor_name", "action", "entity", "entity_id", "detail", "created_at"]
    wb = Workbook()
    ws = wb.active
    ws.title = "AuditLogs"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(key) for key in headers])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
