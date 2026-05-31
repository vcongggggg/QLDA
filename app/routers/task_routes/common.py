from datetime import datetime, timezone
from io import BytesIO, StringIO
import csv

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook

from app.auth import get_current_user, is_member_role, require_permission, require_roles
from app.deps import require_project_access
from app.repository import (
    build_kanban_summary,
    bulk_update_tasks,
    carryover_sprint_tasks,
    create_audit_log,
    create_app_notification,
    create_kanban_saved_filter,
    create_project_milestone,
    create_recurring_task_rule,
    create_task,
    create_task_comment,
    create_task_template,
    delete_kanban_saved_filter,
    delete_task_template,
    duplicate_task,
    get_kanban_saved_filter,
    get_kanban_wip_policy,
    get_project_milestone,
    get_task_ai_detail,
    get_task_by_id,
    get_task_detail_by_id,
    get_task_template,
    get_sprint_by_id,
    get_user_by_username_or_email,
    get_user_by_id,
    list_project_backlog,
    list_project_milestones,
    list_recurring_task_rules,
    list_kanban_saved_filters,
    list_task_activity_logs,
    list_task_comments,
    list_task_templates,
    list_tasks,
    project_exists,
    replace_task_dependencies,
    run_due_recurring_task_rules,
    sprint_exists,
    update_kanban_saved_filter,
    update_project_milestone,
    update_task_deadline,
    update_task_backlog_grooming,
    update_task_milestone,
    update_task_template,
    task_project_ids,
    update_task_metadata,
    update_task_status,
    upsert_kanban_wip_policy,
    user_exists,
)
from app.schemas import (
    BacklogMoveToSprint,
    BacklogGroomingUpdate,
    KanbanSavedFilterCreate,
    KanbanSavedFilterOut,
    KanbanSummaryOut,
    KanbanWipPolicyOut,
    KanbanWipPolicyUpdate,
    MilestoneCreate,
    MilestoneOut,
    MilestoneUpdate,
    RecurringTaskRuleCreate,
    RecurringTaskRuleOut,
    RecurringTaskRunResult,
    SprintCarryoverRequest,
    TaskBulkAction,
    TaskBulkActionResult,
    TaskCommentCreate,
    TaskCommentOut,
    TaskCreate,
    TaskDeadlineExtension,
    TaskDependencyUpdate,
    TaskDetailOut,
    TaskDuplicateRequest,
    TaskFromTemplateRequest,
    TaskImportResult,
    TaskMetadataUpdate,
    TaskMilestoneUpdate,
    TaskOut,
    TaskStatusUpdate,
    TaskTemplateCreate,
    TaskTemplateOut,
)


def _ensure_task_visible(task: dict, current_user: dict) -> None:
    if is_member_role(current_user) and int(task["assignee_id"]) != int(current_user["id"]):
        raise HTTPException(status_code=403, detail="member can access only assigned tasks")


def _ensure_task_project_access(task: dict, current_user: dict) -> None:
    project_id = task.get("project_id")
    if project_id is not None:
        require_project_access(current_user, int(project_id))


def _validate_priority(priority: str | None) -> None:
    if priority is not None and priority not in {"low", "medium", "high", "urgent"}:
        raise HTTPException(status_code=400, detail="priority must be one of low|medium|high|urgent")


def _validate_string_list(name: str, values: list[str] | None, max_items: int = 50, max_len: int = 160) -> None:
    if values is None:
        return
    if len(values) > max_items:
        raise HTTPException(status_code=400, detail=f"{name} has too many items")
    for value in values:
        if not isinstance(value, str) or not value.strip():
            raise HTTPException(status_code=400, detail=f"{name} items must be non-empty strings")
        if len(value) > max_len:
            raise HTTPException(status_code=400, detail=f"{name} items must be at most {max_len} characters")


def _validate_attachment_metadata(items: list[dict] | None) -> None:
    if items is None:
        return
    if len(items) > 20:
        raise HTTPException(status_code=400, detail="attachment_metadata has too many items")
    for item in items:
        if not isinstance(item, dict):
            raise HTTPException(status_code=400, detail="attachment_metadata items must be objects")
        name = str(item.get("name") or "").strip()
        url = str(item.get("url") or item.get("storage_path") or "").strip()
        if not name or len(name) > 200:
            raise HTTPException(status_code=400, detail="attachment_metadata name is required")
        if url and len(url) > 500:
            raise HTTPException(status_code=400, detail="attachment_metadata reference is too long")


def _validate_metadata_payload(payload: TaskCreate | TaskMetadataUpdate) -> None:
    _validate_priority(payload.priority)
    _validate_string_list("labels", payload.labels, max_items=20, max_len=80)
    _validate_string_list("checklist", payload.checklist, max_items=50, max_len=200)
    _validate_string_list("subtasks", payload.subtasks, max_items=50, max_len=200)
    _validate_string_list("dependencies", payload.dependencies, max_items=50, max_len=160)
    _validate_attachment_metadata(payload.attachment_metadata)


def _load_tasks_or_404(task_ids: list[int]) -> list[dict]:
    tasks = []
    missing = []
    for task_id in task_ids:
        task = get_task_by_id(task_id)
        if task:
            tasks.append(task)
        else:
            missing.append(task_id)
    if missing:
        raise HTTPException(status_code=404, detail="task not found")
    return tasks


def _parse_dt(value: str) -> datetime:
    parsed = datetime.fromisoformat(value)
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _query_datetime_to_utc(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _normalize_query_datetime(value: datetime | None) -> str | None:
    normalized = _query_datetime_to_utc(value)
    return normalized.isoformat() if normalized is not None else None


def _due_state(task: dict) -> str:
    deadline = _parse_dt(str(task["deadline"]))
    completed_at = task.get("completed_at")
    if completed_at:
        return "on_time" if _parse_dt(str(completed_at)) <= deadline else "late"
    return "overdue" if deadline < datetime.now(timezone.utc) else "on_time"


def _task_detail_payload(task_id: int) -> dict:
    task = get_task_detail_by_id(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="task not found")
    return {
        **task,
        "due_state": _due_state(task),
        "ai_detail": get_task_ai_detail(task_id),
        "comments": list_task_comments(task_id),
        "activity_logs": list_task_activity_logs(task_id),
    }


def _status_label(status: str) -> str:
    return {"todo": "To-do", "doing": "Doing", "done": "Done"}.get(status, status)


def _ensure_manager_project_access(project_id: int | None, current_user: dict) -> None:
    if project_id is not None:
        require_project_access(current_user, project_id)


def _validate_filters_payload(filters: dict) -> dict:
    allowed = {
        "assignee_id",
        "project_id",
        "sprint_id",
        "status",
        "overdue",
        "keyword",
        "deadline_from",
        "deadline_to",
    }
    clean = {key: value for key, value in (filters or {}).items() if key in allowed and value not in ("", None)}
    if "status" in clean and clean["status"] not in {"todo", "doing", "done"}:
        raise HTTPException(status_code=400, detail="status must be one of todo|doing|done")
    return clean


def _filtered_tasks_for_request(
    current_user: dict,
    assignee_id: int | None = None,
    project_id: int | None = None,
    sprint_id: int | None = None,
    status: str | None = None,
    overdue: bool | None = None,
    keyword: str | None = None,
    deadline_from: datetime | None = None,
    deadline_to: datetime | None = None,
    as_of: datetime | None = None,
) -> list[dict]:
    if is_member_role(current_user):
        assignee_id = int(current_user["id"])
    if status is not None and status not in {"todo", "doing", "done"}:
        raise HTTPException(status_code=400, detail="status must be one of todo|doing|done")
    normalized_deadline_from = _query_datetime_to_utc(deadline_from)
    normalized_deadline_to = _query_datetime_to_utc(deadline_to)
    if (
        normalized_deadline_from is not None
        and normalized_deadline_to is not None
        and normalized_deadline_to < normalized_deadline_from
    ):
        raise HTTPException(status_code=400, detail="deadline_to must be greater than or equal to deadline_from")
    return list_tasks(
        assignee_id=assignee_id,
        project_id=project_id,
        sprint_id=sprint_id,
        status=status,
        overdue=overdue,
        keyword=keyword.strip() if keyword else None,
        deadline_from=_normalize_query_datetime(normalized_deadline_from),
        deadline_to=_normalize_query_datetime(normalized_deadline_to),
        as_of=_normalize_query_datetime(as_of),
    )


def _parse_import_deadline(value: str) -> datetime:
    try:
        parsed = datetime.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise ValueError("deadline must be ISO datetime") from exc
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _parse_labels(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in str(value).split(";") if item.strip()]


def _read_task_import_rows(file: UploadFile, content: bytes) -> list[dict]:
    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(StringIO(text)))
    if filename.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell or "").strip() for cell in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]
    raise HTTPException(status_code=400, detail="only .csv and .xlsx files are supported")


def _validate_template_payload(payload: TaskTemplateCreate) -> None:
    _validate_priority(payload.priority)
    if payload.difficulty not in {"easy", "medium", "hard"}:
        raise HTTPException(status_code=400, detail="difficulty must be one of easy|medium|hard")
    _validate_string_list("labels", payload.labels, max_items=20, max_len=80)
    _validate_string_list("checklist", payload.checklist, max_items=50, max_len=200)
    _validate_string_list("subtasks", payload.subtasks, max_items=50, max_len=200)



__all__ = [name for name in globals() if not name.startswith("__")]
