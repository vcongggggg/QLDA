from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from io import BytesIO, StringIO
import csv
from openpyxl import load_workbook

from app.auth import current_role_code, get_current_user, has_permission, is_member_role, require_permission
from app.kpi import (
    DEFAULT_KPI_POLICY,
    calculate_monthly_kpi_from_transactions,
    compute_dashboard_metrics,
    policy_from_row,
    policy_to_dict,
)
from app.repository import (
    all_tasks_with_users,
    create_audit_log,
    create_kpi_adjustment,
    get_kpi_adjustment,
    get_kpi_policy,
    list_kpi_target_progress,
    list_kpi_targets,
    list_kpi_transactions,
    rebuild_kpi_transactions,
    review_kpi_adjustment,
    save_kpi_policy,
    update_kpi_target,
    upsert_kpi_target,
    user_exists,
)
from app.schemas import (
    DashboardSummary,
    KPIAdjustmentCreate,
    KPIAdjustmentOut,
    KPIAdjustmentReview,
    KPIConfigOut,
    KPIConfigUpdate,
    KPIDepartmentBreakdownOut,
    KPIHistoryRow,
    KPITargetCreate,
    KPITargetImportOut,
    KPITargetOut,
    KPITargetProgressOut,
    KPITargetUpdate,
    KPITeamSummaryOut,
    KPITransactionOut,
    KPITransactionRebuildOut,
    KPIUserResult,
)

router = APIRouter(tags=["kpi"])


def _can_manage_kpi(user: dict) -> bool:
    return has_permission(user, "KPI_MANAGE") or has_permission(user, "kpi.adjust")


def _require_manage_kpi(user: dict) -> None:
    if not _can_manage_kpi(user):
        raise HTTPException(status_code=403, detail="forbidden")


def _current_policy():
    return policy_from_row(get_kpi_policy())


def _config_response(row: dict | None = None) -> dict:
    if row:
        policy = policy_from_row(row)
        data = policy_to_dict(policy)
        data.update(
            {
                "id": row.get("id"),
                "change_reason": row.get("change_reason"),
                "updated_by": row.get("updated_by"),
                "updated_at": row.get("updated_at"),
            }
        )
        return data
    data = policy_to_dict(DEFAULT_KPI_POLICY)
    data.update({"id": None, "change_reason": None, "updated_by": None, "updated_at": None})
    return data


def _monthly_rows(month: str, current_user: dict) -> list[dict]:
    rebuild_kpi_transactions(month, _current_policy())
    user_id = int(current_user["id"]) if is_member_role(current_user) else None
    rows = calculate_monthly_kpi_from_transactions(
        list_kpi_transactions(month, user_id=user_id, include_reversed=False),
        month,
    )
    progress = {int(row["user_id"]): row for row in list_kpi_target_progress(month)}
    for row in rows.values():
        target = progress.get(int(row["user_id"]))
        if target:
            row["target_score"] = target["target_score"]
            row["progress_percent"] = target["progress_percent"]
            row["gap"] = target["gap"]
    return sorted(rows.values(), key=lambda item: item["score"], reverse=True)


def _read_import_rows(file: UploadFile, content: bytes) -> list[dict]:
    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        return list(csv.DictReader(StringIO(content.decode("utf-8-sig"))))
    if filename.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell or "").strip() for cell in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]
    raise HTTPException(status_code=400, detail="only .csv and .xlsx files are supported")


def _month_sequence(end_month: str, months: int) -> list[str]:
    year, month = [int(part) for part in end_month.split("-", 1)]
    values: list[str] = []
    for _ in range(months):
        values.append(f"{year:04d}-{month:02d}")
        month -= 1
        if month == 0:
            year -= 1
            month = 12
    return list(reversed(values))


@router.get("/kpi/config", response_model=KPIConfigOut)
def get_kpi_config_endpoint(current_user: dict = Depends(get_current_user)) -> dict:
    require_permission(current_user, "kpi.view")
    return _config_response(get_kpi_policy())


@router.put("/kpi/config", response_model=KPIConfigOut)
def update_kpi_config_endpoint(payload: KPIConfigUpdate, current_user: dict = Depends(get_current_user)) -> dict:
    _require_manage_kpi(current_user)
    required = {"easy", "medium", "hard"}
    if set(payload.difficulty_multiplier) != required:
        raise HTTPException(status_code=400, detail="difficulty_multiplier must include easy, medium and hard")
    if payload.fallback_difficulty not in required:
        raise HTTPException(status_code=400, detail="fallback_difficulty must be one of easy|medium|hard")
    row = save_kpi_policy(
        difficulty_multiplier={key: float(payload.difficulty_multiplier[key]) for key in sorted(required)},
        on_time_points=payload.on_time_points,
        late_points=payload.late_points,
        overdue_unfinished_points=payload.overdue_unfinished_points,
        fallback_difficulty=payload.fallback_difficulty,
        change_reason=payload.change_reason,
        updated_by=int(current_user["id"]),
    )
    create_audit_log(current_user["id"], "update", "kpi_policy", row["id"], payload.change_reason)
    return _config_response(row)


@router.get("/kpi/monthly", response_model=list[KPIUserResult])
def monthly_kpi_endpoint(
    month: str = Query(description="YYYY-MM"),
    current_user: dict = Depends(get_current_user),
) -> list[dict]:
    require_permission(current_user, "kpi.view")
    return _monthly_rows(month, current_user)


@router.post("/kpi/transactions/rebuild", response_model=KPITransactionRebuildOut)
def rebuild_kpi_transactions_endpoint(
    month: str = Query(description="YYYY-MM"),
    current_user: dict = Depends(get_current_user),
) -> dict:
    _require_manage_kpi(current_user)
    result = rebuild_kpi_transactions(month, _current_policy())
    create_audit_log(current_user["id"], "rebuild", "kpi_transactions", None, f"month={month}")
    return {"month": month, **result}


@router.get("/kpi/transactions", response_model=list[KPITransactionOut])
def list_kpi_transactions_endpoint(
    month: str = Query(description="YYYY-MM"),
    user_id: int | None = Query(default=None),
    current_user: dict = Depends(get_current_user),
) -> list[dict]:
    require_permission(current_user, "kpi.view")
    if is_member_role(current_user):
        user_id = int(current_user["id"])
    rebuild_kpi_transactions(month, _current_policy())
    return list_kpi_transactions(month, user_id=user_id, include_reversed=True)


@router.get("/dashboard/summary", response_model=DashboardSummary)
def dashboard_summary_endpoint(
    month: str = Query(description="YYYY-MM"),
    as_of: datetime | None = Query(default=None),
    current_user: dict = Depends(get_current_user),
) -> dict:
    rebuild_kpi_transactions(month, _current_policy())
    tasks = all_tasks_with_users()
    if is_member_role(current_user):
        tasks = [t for t in tasks if int(t["assignee_id"]) == int(current_user["id"])]
    monthly_kpi = calculate_monthly_kpi_from_transactions(
        list_kpi_transactions(month, user_id=int(current_user["id"]) if is_member_role(current_user) else None, include_reversed=False),
        month,
    )
    normalized_as_of = None
    if as_of is not None:
        normalized_as_of = as_of.replace(tzinfo=timezone.utc) if as_of.tzinfo is None else as_of.astimezone(timezone.utc)
    return compute_dashboard_metrics(tasks, monthly_kpi, month, as_of=normalized_as_of)


@router.post("/kpi/adjustments", response_model=KPIAdjustmentOut)
def create_kpi_adjustment_endpoint(
    payload: KPIAdjustmentCreate,
    current_user: dict = Depends(get_current_user),
) -> dict:
    require_permission(current_user, "kpi.adjust")
    if not user_exists(payload.user_id):
        raise HTTPException(status_code=404, detail="target user not found")
    role = current_role_code(current_user)
    auto_approved = role in {"ADMIN", "HR"}
    now = datetime.now(timezone.utc).isoformat() if auto_approved else None
    item = create_kpi_adjustment(
        user_id=payload.user_id,
        month=payload.month,
        points=payload.points,
        reason=payload.reason,
        created_by=current_user["id"],
        status="approved" if auto_approved else "pending",
        reviewer_id=int(current_user["id"]) if auto_approved else None,
        reviewed_at=now,
        review_reason="auto-approved by privileged KPI role" if auto_approved else None,
    )
    create_audit_log(current_user["id"], "create", "kpi_adjustment", item["id"], payload.reason)
    if auto_approved:
        rebuild_kpi_transactions(payload.month, _current_policy())
    return item


@router.post("/kpi/adjustments/{adjustment_id}/approve", response_model=KPIAdjustmentOut)
def approve_kpi_adjustment_endpoint(
    adjustment_id: int,
    payload: KPIAdjustmentReview,
    current_user: dict = Depends(get_current_user),
) -> dict:
    if current_role_code(current_user) not in {"ADMIN", "HR"}:
        raise HTTPException(status_code=403, detail="forbidden")
    item = review_kpi_adjustment(adjustment_id, "approved", int(current_user["id"]), payload.review_reason)
    if not item:
        raise HTTPException(status_code=404, detail="adjustment not found")
    create_audit_log(current_user["id"], "approve", "kpi_adjustment", adjustment_id, payload.review_reason)
    rebuild_kpi_transactions(str(item["month"]), _current_policy())
    return item


@router.post("/kpi/adjustments/{adjustment_id}/reject", response_model=KPIAdjustmentOut)
def reject_kpi_adjustment_endpoint(
    adjustment_id: int,
    payload: KPIAdjustmentReview,
    current_user: dict = Depends(get_current_user),
) -> dict:
    if current_role_code(current_user) not in {"ADMIN", "HR"}:
        raise HTTPException(status_code=403, detail="forbidden")
    existing = get_kpi_adjustment(adjustment_id)
    if not existing:
        raise HTTPException(status_code=404, detail="adjustment not found")
    item = review_kpi_adjustment(adjustment_id, "rejected", int(current_user["id"]), payload.review_reason)
    create_audit_log(current_user["id"], "reject", "kpi_adjustment", adjustment_id, payload.review_reason)
    rebuild_kpi_transactions(str(existing["month"]), _current_policy())
    return item


@router.get("/kpi/targets", response_model=list[KPITargetOut])
def list_kpi_targets_endpoint(
    month: str | None = Query(default=None),
    user_id: int | None = Query(default=None),
    current_user: dict = Depends(get_current_user),
) -> list[dict]:
    require_permission(current_user, "kpi.view")
    if is_member_role(current_user):
        user_id = int(current_user["id"])
    return list_kpi_targets(month=month, user_id=user_id)


@router.post("/kpi/targets", response_model=KPITargetOut)
def upsert_kpi_target_endpoint(payload: KPITargetCreate, current_user: dict = Depends(get_current_user)) -> dict:
    _require_manage_kpi(current_user)
    if not user_exists(payload.user_id):
        raise HTTPException(status_code=404, detail="target user not found")
    item = upsert_kpi_target(
        user_id=payload.user_id,
        month=payload.month,
        target_score=payload.target_score,
        department_id=payload.department_id,
        team=payload.team,
        created_by=int(current_user["id"]),
    )
    create_audit_log(current_user["id"], "upsert", "kpi_target", item["id"], f"month={payload.month}")
    return item


@router.patch("/kpi/targets/{target_id}", response_model=KPITargetOut)
def update_kpi_target_endpoint(
    target_id: int,
    payload: KPITargetUpdate,
    current_user: dict = Depends(get_current_user),
) -> dict:
    _require_manage_kpi(current_user)
    item = update_kpi_target(target_id, payload.target_score, payload.department_id, payload.team)
    if not item:
        raise HTTPException(status_code=404, detail="target not found")
    create_audit_log(current_user["id"], "update", "kpi_target", target_id, "target updated")
    return item


@router.get("/kpi/targets/progress", response_model=list[KPITargetProgressOut])
def kpi_target_progress_endpoint(
    month: str = Query(description="YYYY-MM"),
    current_user: dict = Depends(get_current_user),
) -> list[dict]:
    require_permission(current_user, "kpi.view")
    rebuild_kpi_transactions(month, _current_policy())
    rows = list_kpi_target_progress(month)
    if is_member_role(current_user):
        rows = [row for row in rows if int(row["user_id"]) == int(current_user["id"])]
    return rows


@router.post("/kpi/targets/import", response_model=KPITargetImportOut)
async def import_kpi_targets_endpoint(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
) -> dict:
    _require_manage_kpi(current_user)
    rows = _read_import_rows(file, await file.read())
    if not rows:
        raise HTTPException(status_code=400, detail="import file has no target rows")
    prepared: list[dict] = []
    errors: list[dict] = []
    for index, row in enumerate(rows, start=2):
        try:
            user_id = int(row.get("user_id") or 0)
            if not user_exists(user_id):
                raise ValueError("user_id not found")
            month = str(row.get("month") or "").strip()
            if len(month) != 7:
                raise ValueError("month must be YYYY-MM")
            prepared.append(
                {
                    "user_id": user_id,
                    "month": month,
                    "target_score": float(row.get("target_score")),
                    "department_id": int(row["department_id"]) if row.get("department_id") not in (None, "") else None,
                    "team": str(row.get("team") or "").strip() or None,
                }
            )
        except (TypeError, ValueError) as exc:
            errors.append({"row": index, "error": str(exc)})
    if errors:
        raise HTTPException(status_code=400, detail={"errors": errors})
    targets = [upsert_kpi_target(created_by=int(current_user["id"]), **item) for item in prepared]
    create_audit_log(current_user["id"], "import", "kpi_target", None, f"count={len(targets)};source={file.filename}")
    return {"upserted_count": len(targets), "targets": targets}


@router.get("/kpi/history", response_model=list[KPIHistoryRow])
def kpi_history_endpoint(
    user_id: int,
    months: int = Query(default=6, ge=1, le=12),
    end_month: str | None = Query(default=None),
    current_user: dict = Depends(get_current_user),
) -> list[dict]:
    require_permission(current_user, "kpi.view")
    if is_member_role(current_user) and int(current_user["id"]) != user_id:
        raise HTTPException(status_code=403, detail="forbidden")
    end = end_month or datetime.now(timezone.utc).strftime("%Y-%m")
    out: list[dict] = []
    for month in _month_sequence(end, months):
        rows = _monthly_rows(month, {"id": user_id, "role": "MEMBER", "role_id": "MEMBER"})
        if rows:
            out.append(rows[0])
        else:
            out.append({"user_id": user_id, "user_name": "Unknown", "month": month, "score": 0.0})
    return out


@router.get("/kpi/team-summary", response_model=KPITeamSummaryOut)
def kpi_team_summary_endpoint(month: str, current_user: dict = Depends(get_current_user)) -> dict:
    require_permission(current_user, "kpi.view")
    rows = _monthly_rows(month, current_user)
    targets = [row for row in rows if row.get("target_score") is not None]
    scores = [float(row["score"]) for row in rows]
    return {
        "month": month,
        "user_count": len(rows),
        "avg_score": round(sum(scores) / len(scores), 2) if scores else 0.0,
        "avg_target": round(sum(float(row["target_score"]) for row in targets) / len(targets), 2) if targets else None,
        "below_target_count": sum(1 for row in rows if row.get("gap", 0) and float(row.get("gap") or 0) > 0),
    }


@router.get("/kpi/department-breakdown", response_model=list[KPIDepartmentBreakdownOut])
def kpi_department_breakdown_endpoint(month: str, current_user: dict = Depends(get_current_user)) -> list[dict]:
    require_permission(current_user, "kpi.view")
    progress = list_kpi_target_progress(month)
    groups: dict[str, dict] = {}
    for row in progress:
        key = str(row.get("department_id") or "none")
        group = groups.setdefault(
            key,
            {
                "department_id": row.get("department_id"),
                "department_name": row.get("department_name") or "Unassigned",
                "month": month,
                "scores": [],
                "targets": [],
                "below_target_count": 0,
            },
        )
        group["scores"].append(float(row["score"]))
        group["targets"].append(float(row["target_score"]))
        if float(row["gap"]) > 0:
            group["below_target_count"] += 1
    return [
        {
            "department_id": group["department_id"],
            "department_name": group["department_name"],
            "month": month,
            "user_count": len(group["scores"]),
            "avg_score": round(sum(group["scores"]) / len(group["scores"]), 2) if group["scores"] else 0.0,
            "avg_target": round(sum(group["targets"]) / len(group["targets"]), 2) if group["targets"] else None,
            "below_target_count": group["below_target_count"],
        }
        for group in groups.values()
    ]
